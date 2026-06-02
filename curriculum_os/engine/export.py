import json
import re
import hashlib
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

from curriculum_os.engine.models import CurriculumExtraction


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")[:48] or "item"

def _stable_id(*parts: str) -> str:
    """
    Deterministic ID for stable exports.
    Keeps IDs consistent across reruns as long as inputs don't change.
    """
    raw = "||".join(p.strip() for p in parts if p is not None)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return digest


def build_knowledge_graph(
    documents: list[dict],
) -> dict:
    """
    Build a hybrid relational + graph representation.

    Each document dict: grade, source_file, extractions (list of dicts).
    """
    publisher = "Hong Kong Metropolitan University"

    units: list[dict] = []
    grammar_points: list[dict] = []
    vocabulary_themes: list[dict] = []
    vocabulary_words: list[dict] = []
    skills: list[dict] = []
    edges: list[dict] = []

    unit_index: dict[tuple[str, str, str], str] = {}

    for doc in documents:
        grade = doc["grade"]
        source = doc["source_file"]
        textbook = Path(source).stem

        for i, raw in enumerate(doc["extractions"]):
            ext = (
                raw
                if isinstance(raw, CurriculumExtraction)
                else CurriculumExtraction.model_validate(raw)
            )
            title = (ext.unit_title or f"Chunk {i + 1}").strip()
            # unit uniqueness: (grade, source_file, title)
            unit_key = (grade, source, _slug(title))
            if unit_key not in unit_index:
                unit_id = f"U{_stable_id(grade, source, title)}"
                unit_index[unit_key] = unit_id
                units.append(
                    {
                        "unit_id": unit_id,
                        "grade": grade,
                        "source_file": source,
                        "publisher": publisher,
                        "textbook": textbook,
                        "unit_title": title,
                    }
                )
            unit_id = unit_index[unit_key]

            for g_idx, grammar in enumerate(ext.grammar_points):
                g = grammar.strip()
                if not g:
                    continue
                grammar_id = f"G{_stable_id(unit_id, 'grammar', str(g_idx), g)}"
                grammar_points.append(
                    {
                        "grammar_id": grammar_id,
                        "unit_id": unit_id,
                        "grammar_point": g,
                    }
                )
                edges.append(
                    {
                        "source_id": unit_id,
                        "target_id": grammar_id,
                        "relation_type": "teaches",
                        "source_type": "Unit",
                        "target_type": "GrammarPoint",
                    }
                )

            for v_idx, vocab in enumerate(ext.vocabulary_themes):
                v = vocab.strip()
                if not v:
                    continue
                vocab_theme_id = f"VT{_stable_id(unit_id, 'theme', str(v_idx), v)}"
                vocabulary_themes.append(
                    {
                        "vocab_theme_id": vocab_theme_id,
                        "unit_id": unit_id,
                        "theme": v,
                    }
                )
                edges.append(
                    {
                        "source_id": unit_id,
                        "target_id": vocab_theme_id,
                        "relation_type": "covers_theme",
                        "source_type": "Unit",
                        "target_type": "VocabularyTheme",
                    }
                )

            for w_idx, word in enumerate(getattr(ext, "vocabulary_words", []) or []):
                w = str(word).strip()
                if not w:
                    continue
                vocab_word_id = f"VW{_stable_id(unit_id, 'word', str(w_idx), w)}"
                vocabulary_words.append(
                    {
                        "vocab_word_id": vocab_word_id,
                        "unit_id": unit_id,
                        "word": w,
                    }
                )
                edges.append(
                    {
                        "source_id": unit_id,
                        "target_id": vocab_word_id,
                        "relation_type": "teaches_word",
                        "source_type": "Unit",
                        "target_type": "VocabularyWord",
                    }
                )

            for s_idx, skill in enumerate(ext.language_skills):
                s = skill.strip()
                if not s:
                    continue
                skill_id = f"S{_stable_id(unit_id, 'skill', str(s_idx), s)}"
                skills.append(
                    {
                        "skill_id": skill_id,
                        "unit_id": unit_id,
                        "language_skill": s,
                    }
                )
                edges.append(
                    {
                        "source_id": unit_id,
                        "target_id": skill_id,
                        "relation_type": "develops",
                        "source_type": "Unit",
                        "target_type": "Skill",
                    }
                )

    return {
        "metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "grades_processed": sorted({d["grade"] for d in documents}),
            "document_count": len(documents),
            "edge_count": len(edges),
            "documents": [
                {
                    "grade": d["grade"],
                    "source_file": d["source_file"],
                    "extraction_count": len(d["extractions"]),
                    "document_type": d.get("document_analysis", {})
                    .get("document_type", {})
                    .get("document_type", "unknown"),
                    "level_confidence": d.get("document_analysis", {}).get("level_confidence", 0.0),
                }
                for d in documents
            ],
        },
        "tables": {
            "Units": units,
            "GrammarPoints": grammar_points,
            "VocabularyThemes": vocabulary_themes,
            "VocabularyWords": vocabulary_words,
            "Skills": skills,
            "Edges": edges,
        },
        # Keep raw per-document results for audit/debug.
        "extractions_by_document": [
            {
                "grade": d["grade"],
                "source_file": d["source_file"],
                "document_analysis": d.get("document_analysis", {}),
                "units": d["extractions"],
            }
            for d in documents
        ],
    }


def export_to_json(graph: dict, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(graph, f, indent=2, ensure_ascii=False)
    return path


def export_to_excel(graph: dict, path: str | Path) -> Path:
    """Export relational tables into workbook sheets."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    tables = graph.get("tables", {})

    units_ws = wb.active
    units_ws.title = "Units"
    units_ws.append(["unit_id", "grade", "source_file", "publisher", "textbook", "unit_title"])
    for row in tables.get("Units", []):
        units_ws.append(
            [
                row.get("unit_id", ""),
                row.get("grade", ""),
                row.get("source_file", ""),
                row.get("publisher", ""),
                row.get("textbook", ""),
                row.get("unit_title", ""),
            ]
        )

    grammar_ws = wb.create_sheet("GrammarPoints")
    grammar_ws.append(["grammar_id", "unit_id", "grammar_point"])
    for row in tables.get("GrammarPoints", []):
        grammar_ws.append([row.get("grammar_id", ""), row.get("unit_id", ""), row.get("grammar_point", "")])

    vocab_t_ws = wb.create_sheet("VocabularyThemes")
    vocab_t_ws.append(["vocab_theme_id", "unit_id", "theme"])
    for row in tables.get("VocabularyThemes", []):
        vocab_t_ws.append([row.get("vocab_theme_id", ""), row.get("unit_id", ""), row.get("theme", "")])

    vocab_w_ws = wb.create_sheet("VocabularyWords")
    vocab_w_ws.append(["vocab_word_id", "unit_id", "word"])
    for row in tables.get("VocabularyWords", []):
        vocab_w_ws.append([row.get("vocab_word_id", ""), row.get("unit_id", ""), row.get("word", "")])

    skills_ws = wb.create_sheet("Skills")
    skills_ws.append(["skill_id", "unit_id", "language_skill"])
    for row in tables.get("Skills", []):
        skills_ws.append([row.get("skill_id", ""), row.get("unit_id", ""), row.get("language_skill", "")])

    edges_ws = wb.create_sheet("Edges")
    edges_ws.append(["source_id", "target_id", "relation_type", "source_type", "target_type"])
    for row in tables.get("Edges", []):
        edges_ws.append(
            [
                row.get("source_id", ""),
                row.get("target_id", ""),
                row.get("relation_type", ""),
                row.get("source_type", ""),
                row.get("target_type", ""),
            ]
        )

    wb.save(path)
    return path
